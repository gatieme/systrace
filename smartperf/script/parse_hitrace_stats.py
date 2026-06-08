#!/usr/bin/env python3
"""Read HiSmartPerf .db traces, merge by filename timestamp, export 5s CPU/GPU stats to xlsx."""

from __future__ import annotations

import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

LOCAL_TZ = ZoneInfo("Asia/Shanghai")
TRACE_TS_RE = re.compile(r"record_trace_(\d{14})@|trace_(\d{14})@")
EXPECTED_CPUS = list(range(16))
CPU_GROUP_INTERVAL_SEC = 5
LINE_COLORS = (
    "4472C4",
    "ED7D31",
    "70AD47",
    "FFC000",
    "5B9BD5",
    "A5A5A5",
    "264478",
    "9E480E",
    "636363",
    "997300",
    "C55A11",
    "843C0C",
    "375623",
    "7030A0",
    "C00000",
    "00B0F0",
)
CPU_GROUP_RANGES: tuple[tuple[str, range], ...] = (
    ("CPU0-3", range(0, 4)),
    ("CPU4-9", range(4, 10)),
    ("CPU10-15", range(10, 16)),
)
GROUP_LINE_COLORS = ("4472C4", "ED7D31", "70AD47")
GPU_LINE_COLORS = ("7030A0", "C00000")
CPU_GPU_MERGE_COLORS = GROUP_LINE_COLORS + (GPU_LINE_COLORS[0],)
GPU_MEASURE_NAMES = ("gpuload", "gpufreq")
HZ_TO_MHZ = 1_000_000.0


def trace_file_sort_key(path: Path) -> tuple[int, str]:
    match = TRACE_TS_RE.search(path.name)
    if match:
        ts = match.group(1) or match.group(2)
        return int(ts), path.name
    return 0, path.name


def filename_wall_base_ns(path: Path) -> int:
    match = TRACE_TS_RE.search(path.name)
    if not match:
        raise ValueError(f"Cannot parse trace timestamp from filename: {path.name}")
    ts = match.group(1) or match.group(2)
    dt = datetime.strptime(ts, "%Y%m%d%H%M%S").replace(tzinfo=LOCAL_TZ)
    return int(dt.timestamp() * 1_000_000_000)


def load_running_intervals(db_path: Path) -> tuple[int, int, list[dict]]:
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute("SELECT start_ts, end_ts FROM trace_range").fetchone()
        if not row:
            return 0, 0, []
        trace_start_ts, trace_end_ts = int(row[0]), int(row[1])
        wall_base_ns = filename_wall_base_ns(db_path)
        query = """
            SELECT ts, dur, cpu
            FROM thread_state
            WHERE state = 'Running'
              AND cpu IS NOT NULL
              AND dur > 0
              AND (ts - ?) > 0
        """
        intervals = []
        for ts, dur, cpu in conn.execute(query, (trace_start_ts,)):
            cpu_id = int(cpu)
            if cpu_id not in EXPECTED_CPUS:
                continue
            rel_start = int(ts) - trace_start_ts
            intervals.append(
                {
                    "wall_ts_ns": wall_base_ns + rel_start,
                    "dur_ns": int(dur),
                    "cpu_id": cpu_id,
                    "source_file": db_path.name,
                }
            )
        return trace_start_ts, trace_end_ts, intervals
    finally:
        conn.close()


def merge_db_intervals(db_files: list[Path]) -> tuple[list[dict], dict]:
    all_intervals: list[dict] = []
    meta_rows = []
    for db_path in db_files:
        trace_start_ts, trace_end_ts, intervals = load_running_intervals(db_path)
        duration_sec = (trace_end_ts - trace_start_ts) / 1e9 if trace_end_ts > trace_start_ts else 0.0
        meta_rows.append(
            {
                "source_file": db_path.name,
                "trace_start_ts": trace_start_ts,
                "trace_end_ts": trace_end_ts,
                "trace_duration_sec": round(duration_sec, 3),
                "running_slices": len(intervals),
                "wall_base_local": str(
                    datetime.fromtimestamp(
                        filename_wall_base_ns(db_path) / 1e9, tz=LOCAL_TZ
                    )
                ),
            }
        )
        all_intervals.extend(intervals)
    all_intervals.sort(key=lambda item: (item["wall_ts_ns"], item["cpu_id"]))
    return all_intervals, {"files": meta_rows}


def load_gpu_measure_samples(db_path: Path) -> tuple[int, int, list[dict]]:
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute("SELECT start_ts, end_ts FROM trace_range").fetchone()
        if not row:
            return 0, 0, []
        trace_start_ts, trace_end_ts = int(row[0]), int(row[1])
        wall_base_ns = filename_wall_base_ns(db_path)
        trace_end_wall_ns = wall_base_ns + (trace_end_ts - trace_start_ts)

        filter_rows = conn.execute(
            """
            SELECT id, name
            FROM measure_filter
            WHERE name IN ({})
            """.format(",".join("?" for _ in GPU_MEASURE_NAMES)),
            GPU_MEASURE_NAMES,
        ).fetchall()
        if not filter_rows:
            return trace_start_ts, trace_end_ts, []

        filter_name = {int(fid): name for fid, name in filter_rows}
        filter_ids = list(filter_name)
        query = """
            SELECT ts, dur, value, filter_id
            FROM measure
            WHERE filter_id IN ({})
              AND ts > ?
            ORDER BY ts
        """.format(",".join("?" for _ in filter_ids))

        samples = []
        for ts, dur, value, filter_id in conn.execute(
            query, (*filter_ids, trace_start_ts)
        ):
            metric = filter_name.get(int(filter_id))
            if metric is None:
                continue
            rel_start = int(ts) - trace_start_ts
            dur_ns = int(dur) if dur is not None else None
            samples.append(
                {
                    "wall_ts_ns": wall_base_ns + rel_start,
                    "dur_ns": dur_ns,
                    "value": float(value),
                    "metric": metric,
                    "filter_id": int(filter_id),
                    "trace_end_wall_ns": trace_end_wall_ns,
                    "source_file": db_path.name,
                }
            )
        return trace_start_ts, trace_end_ts, samples
    finally:
        conn.close()


def merge_gpu_samples(db_files: list[Path]) -> tuple[list[dict], dict]:
    all_samples: list[dict] = []
    meta_rows = []
    for db_path in db_files:
        trace_start_ts, trace_end_ts, samples = load_gpu_measure_samples(db_path)
        duration_sec = (trace_end_ts - trace_start_ts) / 1e9 if trace_end_ts > trace_start_ts else 0.0
        meta_rows.append(
            {
                "source_file": db_path.name,
                "trace_start_ts": trace_start_ts,
                "trace_end_ts": trace_end_ts,
                "trace_duration_sec": round(duration_sec, 3),
                "gpu_measure_samples": len(samples),
                "wall_base_local": str(
                    datetime.fromtimestamp(
                        filename_wall_base_ns(db_path) / 1e9, tz=LOCAL_TZ
                    )
                ),
            }
        )
        all_samples.extend(samples)
    all_samples.sort(key=lambda item: (item["wall_ts_ns"], item["metric"], item["filter_id"]))
    return all_samples, {"files": meta_rows}


def _sample_end_ns(sample: dict) -> int:
    start = sample["wall_ts_ns"]
    dur_ns = sample["dur_ns"]
    if dur_ns is not None and dur_ns > 0:
        return start + dur_ns
    return int(sample["trace_end_wall_ns"])


def measure_time_weighted_avg(
    samples: list[dict],
    win_start_ns: int,
    win_end_ns: int,
) -> float:
    """Time-weighted average of piecewise-constant measure samples in a window."""
    window_ns = win_end_ns - win_start_ns
    if window_ns <= 0:
        return 0.0

    weighted_sum = 0.0
    for sample in samples:
        start = sample["wall_ts_ns"]
        end = _sample_end_ns(sample)
        if end <= win_start_ns or start >= win_end_ns:
            continue
        clip_start = max(start, win_start_ns)
        clip_end = min(end, win_end_ns)
        if clip_end > clip_start:
            weighted_sum += sample["value"] * (clip_end - clip_start)
    return weighted_sum / window_ns


def gpu_metrics_for_window(
    samples: list[dict],
    win_start_ns: int,
    win_end_ns: int,
) -> dict[str, float]:
    load_samples = [s for s in samples if s["metric"] == "gpuload"]
    freq_samples = [s for s in samples if s["metric"] == "gpufreq"]

    gpuload = measure_time_weighted_avg(load_samples, win_start_ns, win_end_ns)

    freq_by_filter: dict[int, list[dict]] = defaultdict(list)
    for sample in freq_samples:
        freq_by_filter[sample["filter_id"]].append(sample)
    freq_avgs = [
        measure_time_weighted_avg(filter_samples, win_start_ns, win_end_ns)
        for filter_samples in freq_by_filter.values()
    ]
    gpufreq_hz = max(freq_avgs) if freq_avgs else 0.0

    return {
        "gpuload": round(gpuload, 2),
        "gpufreq_mhz": round(gpufreq_hz / HZ_TO_MHZ, 2),
    }


def compute_gpu_measure_5s(samples: list[dict]) -> pd.DataFrame:
    if not samples:
        return pd.DataFrame()

    timeline_start = min(sample["wall_ts_ns"] for sample in samples)
    timeline_end = max(_sample_end_ns(sample) for sample in samples)
    window_ns = CPU_GROUP_INTERVAL_SEC * 1_000_000_000

    rows = []
    win_start = timeline_start
    while win_start < timeline_end:
        win_end = min(win_start + window_ns, timeline_end)
        metrics = gpu_metrics_for_window(samples, win_start, win_end)
        row = {
            "window_start_ns": win_start,
            "window_end_ns": win_end,
            "elapsed_sec": len(rows) * CPU_GROUP_INTERVAL_SEC,
            "gpuload": metrics["gpuload"],
            "gpufreq_mhz": metrics["gpufreq_mhz"],
        }
        rows.append(row)
        win_start = win_end

    return pd.DataFrame(rows)


def build_chart_gpu_dataframe(gpu_df: pd.DataFrame) -> pd.DataFrame:
    if gpu_df.empty:
        return pd.DataFrame()

    x_label, _, x_values, _ = _chart_x_column(gpu_df)
    if not x_label:
        return pd.DataFrame()

    return pd.DataFrame(
        {
            x_label: x_values.values,
            "纵坐标_gpuload(%)": gpu_df["gpuload"].values,
            "纵坐标_gpufreq(MHz)": gpu_df["gpufreq_mhz"].values,
        }
    )


def build_chart_cpu_gpu_dataframe(
    cpu_df: pd.DataFrame,
    gpu_df: pd.DataFrame | None,
) -> pd.DataFrame:
    """Merge chart_xy_group CPU groups and chart_gpu gpuload on aligned 5s windows."""
    has_cpu = not cpu_df.empty
    has_gpu = gpu_df is not None and not gpu_df.empty
    if not has_cpu and not has_gpu:
        return pd.DataFrame()

    elapsed_values: set[int] = set()
    if has_cpu:
        elapsed_values.update(int(v) for v in cpu_df["elapsed_sec"].tolist())
    if has_gpu:
        elapsed_values.update(int(v) for v in gpu_df["elapsed_sec"].tolist())

    rows: list[dict[str, float | int]] = []
    for elapsed in sorted(elapsed_values):
        row_data: dict[str, float | int] = {"横坐标_相对时间(秒)": elapsed}
        if has_cpu:
            cpu_rows = cpu_df[cpu_df["elapsed_sec"] == elapsed]
            for group_name, cpu_range in CPU_GROUP_RANGES:
                if cpu_rows.empty:
                    avg = 0.0
                else:
                    cpu_row = cpu_rows.iloc[0]
                    avg = round(
                        sum(
                            float(cpu_row.get(f"CPU{cpu_id}", 0.0) or 0.0)
                            for cpu_id in cpu_range
                        )
                        / len(cpu_range),
                        2,
                    )
                row_data[f"纵坐标_{group_name}"] = avg
        if has_gpu:
            gpu_rows = gpu_df[gpu_df["elapsed_sec"] == elapsed]
            row_data["纵坐标_gpuload(%)"] = (
                float(gpu_rows.iloc[0]["gpuload"]) if not gpu_rows.empty else 0.0
            )
        rows.append(row_data)

    return pd.DataFrame(rows)


def cpu_usage_for_window(
    intervals: list[dict],
    win_start_ns: int,
    win_end_ns: int,
    cpus: list[int] | None = None,
) -> dict[int, float]:
    """HiSmartPerf getTabCpuUsage: sum clipped Running dur / window length."""
    target_cpus = cpus if cpus is not None else EXPECTED_CPUS
    window_ns = win_end_ns - win_start_ns
    if window_ns <= 0:
        return {cpu_id: 0.0 for cpu_id in target_cpus}

    busy_ns: dict[int, int] = defaultdict(int)
    for row in intervals:
        cpu_id = row["cpu_id"]
        if cpu_id not in target_cpus:
            continue
        start = row["wall_ts_ns"]
        end = start + row["dur_ns"]
        if end <= win_start_ns or start >= win_end_ns:
            continue
        clip_start = max(start, win_start_ns)
        clip_end = min(end, win_end_ns)
        if clip_end > clip_start:
            busy_ns[cpu_id] += clip_end - clip_start

    return {
        cpu_id: round(min(100.0, busy_ns.get(cpu_id, 0) / window_ns * 100), 2)
        for cpu_id in target_cpus
    }


def compute_cpu_usage_5s(intervals: list[dict]) -> pd.DataFrame:
    if not intervals:
        return pd.DataFrame()

    timeline_start = min(row["wall_ts_ns"] for row in intervals)
    timeline_end = max(row["wall_ts_ns"] + row["dur_ns"] for row in intervals)
    window_ns = CPU_GROUP_INTERVAL_SEC * 1_000_000_000

    rows = []
    win_start = timeline_start
    while win_start < timeline_end:
        win_end = min(win_start + window_ns, timeline_end)
        usage = cpu_usage_for_window(intervals, win_start, win_end)
        window_dt_utc = datetime.fromtimestamp(win_start / 1e9, tz=timezone.utc)
        row = {
            "window_start_ns": win_start,
            "window_end_ns": win_end,
            "datetime_utc": window_dt_utc,
            "window_time_local": window_dt_utc.astimezone(LOCAL_TZ).replace(tzinfo=None),
            "elapsed_sec": len(rows) * CPU_GROUP_INTERVAL_SEC,
        }
        for cpu_id in EXPECTED_CPUS:
            row[f"CPU{cpu_id}"] = usage.get(cpu_id, 0.0)
        rows.append(row)
        win_start = win_end

    return pd.DataFrame(rows)


def format_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    export_df = df.copy()
    if "datetime_utc" in export_df.columns:
        export_df["datetime_utc"] = export_df["datetime_utc"].astype(str)
    if "window_time_local" in export_df.columns:
        return export_df
    if "datetime_utc" in export_df.columns:
        utc_series = pd.to_datetime(export_df["datetime_utc"], utc=True, errors="coerce")
        export_df["window_time_local"] = utc_series.dt.tz_convert(LOCAL_TZ).dt.tz_localize(None)
    return export_df


def _chart_x_column(group_df: pd.DataFrame) -> tuple[str, str, pd.Series, bool]:
    """Return (x_label, x_num_fmt, x_values, use_datetime)."""
    if "elapsed_sec" in group_df.columns:
        x_label = "横坐标_相对时间(秒)"
        x_values = group_df["elapsed_sec"]
        return x_label, "0", x_values, False
    return "", "0", pd.Series(dtype=float), False


def _cpu_series(group_df: pd.DataFrame, cpu_id: int) -> pd.Series:
    col = f"CPU{cpu_id}"
    if col in group_df.columns:
        return group_df[col].fillna(0.0)
    return pd.Series(0.0, index=group_df.index)


def build_chart_xy_dataframe(group_df: pd.DataFrame) -> pd.DataFrame:
    if group_df.empty:
        return pd.DataFrame()

    x_label, _, x_values, _ = _chart_x_column(group_df)
    if not x_label:
        return pd.DataFrame()

    chart_df = pd.DataFrame({x_label: x_values.values})
    for cpu_id in EXPECTED_CPUS:
        chart_df[f"纵坐标_CPU{cpu_id}"] = _cpu_series(group_df, cpu_id).values
    return chart_df


def build_chart_xy_group_dataframe(group_df: pd.DataFrame) -> pd.DataFrame:
    """Average CPU0-3 / CPU4-9 / CPU10-15; missing CPUs count as 0."""
    if group_df.empty:
        return pd.DataFrame()

    x_label, _, x_values, _ = _chart_x_column(group_df)
    if not x_label:
        return pd.DataFrame()

    chart_df = pd.DataFrame({x_label: x_values.values})
    for group_name, cpu_range in CPU_GROUP_RANGES:
        group_cols = pd.concat(
            [_cpu_series(group_df, cpu_id) for cpu_id in cpu_range],
            axis=1,
        )
        chart_df[f"纵坐标_{group_name}"] = group_cols.mean(axis=1).round(2).values
    return chart_df


def _style_chart_series(ser: Series, color: str, *, show_data_labels: bool) -> None:
    ser.marker.symbol = "none"
    ser.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill=color, w=20000))
    if show_data_labels:
        ser.dLbls = DataLabelList()
        ser.dLbls.showVal = True
        ser.dLbls.showLegendKey = False
        ser.dLbls.showCatName = False
        ser.dLbls.showSerName = False
        ser.dLbls.numFmt = "0.0"


def _configure_chart_axes(
    chart: ScatterChart,
    *,
    x_axis_title: str,
    y_axis_title: str,
    x_num_fmt: str,
    y_num_fmt: str,
    use_datetime: bool,
    x_max_value: float,
    y_min: float,
    y_max: float,
) -> None:
    chart.x_axis.title = x_axis_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.numFmt = x_num_fmt
    chart.y_axis.numFmt = y_num_fmt
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.majorTickMark = "out"
    chart.y_axis.majorTickMark = "out"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max
    if not use_datetime:
        x_max = max(x_max_value, CPU_GROUP_INTERVAL_SEC)
        x_max = (
            int(x_max + CPU_GROUP_INTERVAL_SEC - 1) // CPU_GROUP_INTERVAL_SEC
        ) * CPU_GROUP_INTERVAL_SEC
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = x_max
        chart.x_axis.majorUnit = CPU_GROUP_INTERVAL_SEC


def _add_scatter_line_chart(
    ws,
    chart_xy_df: pd.DataFrame,
    *,
    chart_title: str,
    colors: tuple[str, ...],
    show_data_labels: bool,
    y_axis_title: str = "CPU Usage (%)",
    y_num_fmt: str = "0.0",
    y_max_cap: float | None = 100.0,
    chart_anchor: str | None = None,
) -> None:
    if chart_xy_df.empty:
        return

    n_rows = len(chart_xy_df) + 1
    columns = list(chart_xy_df.columns)
    x_col = 1
    x_axis_title = columns[0]
    use_datetime = False
    x_num_fmt = "0"

    value_cols = [col for col in columns if col.startswith("纵坐标_")]
    if not value_cols:
        return

    y_min = float(chart_xy_df[value_cols].min().min())
    y_max = float(chart_xy_df[value_cols].max().max())
    y_min = max(0.0, y_min - 2.0)
    y_max = y_max + 2.0
    if y_max_cap is not None:
        y_max = min(y_max_cap, y_max)

    x_max_value = float(chart_xy_df[columns[0]].max())

    chart = ScatterChart()
    chart.title = chart_title
    chart.scatterStyle = "line"
    chart.height = 14
    chart.width = 28
    chart.varyColors = False
    _configure_chart_axes(
        chart,
        x_axis_title=x_axis_title,
        y_axis_title=y_axis_title,
        x_num_fmt=x_num_fmt,
        y_num_fmt=y_num_fmt,
        use_datetime=use_datetime,
        x_max_value=x_max_value,
        y_min=y_min,
        y_max=y_max,
    )

    xvalues = Reference(ws, min_col=x_col, min_row=2, max_row=n_rows)
    for idx, col_name in enumerate(value_cols):
        value_col = columns.index(col_name) + 1
        yvalues = Reference(ws, min_col=value_col, min_row=2, max_row=n_rows)
        series_title = col_name.removeprefix("纵坐标_")
        ser = Series(yvalues, xvalues, title=series_title)
        _style_chart_series(
            ser,
            colors[idx % len(colors)],
            show_data_labels=show_data_labels,
        )
        chart.series.append(ser)

    ws.add_chart(chart, chart_anchor or f"A{n_rows + 3}")


def _add_cpu_charts(
    xlsx_path: Path,
    chart_xy_df: pd.DataFrame,
    chart_xy_group_df: pd.DataFrame,
) -> None:
    if chart_xy_df.empty and chart_xy_group_df.empty:
        return

    wb = load_workbook(xlsx_path)
    if not chart_xy_df.empty and "chart_xy" in wb.sheetnames:
        _add_scatter_line_chart(
            wb["chart_xy"],
            chart_xy_df,
            chart_title="CPU Usage (5s, HiSmartPerf thread_state)",
            colors=LINE_COLORS,
            show_data_labels=True,
        )
    if not chart_xy_group_df.empty and "chart_xy_group" in wb.sheetnames:
        _add_scatter_line_chart(
            wb["chart_xy_group"],
            chart_xy_group_df,
            chart_title="CPU Group Avg (CPU0-3 / CPU4-9 / CPU10-15, 5s)",
            colors=GROUP_LINE_COLORS,
            show_data_labels=True,
        )
    wb.save(xlsx_path)


def _add_gpu_charts(xlsx_path: Path, chart_gpu_df: pd.DataFrame) -> None:
    if chart_gpu_df.empty:
        return

    wb = load_workbook(xlsx_path)
    if "chart_gpu" not in wb.sheetnames:
        wb.save(xlsx_path)
        return

    ws = wb["chart_gpu"]
    n_rows = len(chart_gpu_df) + 1

    load_df = chart_gpu_df[[chart_gpu_df.columns[0], "纵坐标_gpuload(%)"]]
    _add_scatter_line_chart(
        ws,
        load_df,
        chart_title="GPU Load (5s avg, measure gpuload)",
        colors=(GPU_LINE_COLORS[0],),
        show_data_labels=True,
        y_axis_title="GPU Load (%)",
        y_max_cap=100.0,
        chart_anchor=f"A{n_rows + 3}",
    )

    freq_df = chart_gpu_df[[chart_gpu_df.columns[0], "纵坐标_gpufreq(MHz)"]]
    _add_scatter_line_chart(
        ws,
        freq_df,
        chart_title="GPU Frequency (5s avg, measure gpufreq)",
        colors=(GPU_LINE_COLORS[1],),
        show_data_labels=True,
        y_axis_title="GPU Frequency (MHz)",
        y_max_cap=None,
        chart_anchor=f"M{n_rows + 3}",
    )

    wb.save(xlsx_path)


def _add_cpu_gpu_merged_chart(xlsx_path: Path, chart_cpu_gpu_df: pd.DataFrame) -> None:
    if chart_cpu_gpu_df.empty:
        return

    wb = load_workbook(xlsx_path)
    if "chart_cpu_gpu" not in wb.sheetnames:
        wb.save(xlsx_path)
        return

    _add_scatter_line_chart(
        wb["chart_cpu_gpu"],
        chart_cpu_gpu_df,
        chart_title="CPU Group & GPU Load (5s, %)",
        colors=CPU_GPU_MERGE_COLORS,
        show_data_labels=True,
        y_axis_title="Load (%)",
        y_max_cap=100.0,
    )
    wb.save(xlsx_path)


def parse_db_inputs(input_path: Path) -> list[Path]:
    if input_path.is_dir():
        db_files = sorted(input_path.glob("*.db"), key=trace_file_sort_key)
        if not db_files:
            raise FileNotFoundError(f"No .db files found in {input_path}")
        return db_files
    if input_path.suffix.lower() != ".db":
        raise ValueError(f"Expected .db file or directory: {input_path}")
    return [input_path]


def export_xlsx(
    cpu_df: pd.DataFrame,
    file_meta: dict,
    output_path: Path,
    *,
    source_label: str,
    gpu_df: pd.DataFrame | None = None,
) -> None:
    merged_duration = 0.0
    if not cpu_df.empty and "elapsed_sec" in cpu_df.columns:
        merged_duration = float(cpu_df["elapsed_sec"].max()) + CPU_GROUP_INTERVAL_SEC

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary = pd.DataFrame(
            [
                {
                    "source": source_label,
                    "db_file_count": len(file_meta.get("files", [])),
                    "trace_duration_sec": round(merged_duration, 3),
                    "cpu_usage_5s_rows": len(cpu_df),
                    "gpu_measure_5s_rows": 0 if gpu_df is None else len(gpu_df),
                    "cpu_usage_note": (
                        "5s 窗口；HiSmartPerf thread_state Running + trace_range 口径；"
                        "多文件按文件名时间戳墙钟拼接"
                    ),
                    "gpu_measure_note": (
                        "5s 窗口；measure 表 gpuload/gpufreq 时间加权平均；"
                        "gpufreq 取各 clock 域窗口均值的最大值(MHz)"
                    ),
                    "chart_xy_note": (
                        "chart_xy：16 核折线图；chart_xy_group：CPU 三组均值；"
                        "chart_gpu：gpuload/gpufreq；"
                        "chart_cpu_gpu：CPU0-3/4-9/10-15 三组均值 + gpuload 合并折线图"
                    ),
                }
            ]
        )
        summary.to_excel(writer, sheet_name="summary", index=False)

        if file_meta.get("files"):
            pd.DataFrame(file_meta["files"]).to_excel(writer, sheet_name="db_files", index=False)

        chart_xy_df = pd.DataFrame()
        chart_xy_group_df = pd.DataFrame()
        sorted_cpu_df = pd.DataFrame()
        if cpu_df.empty:
            pd.DataFrame({"info": ["未找到 Running 状态 CPU 数据"]}).to_excel(
                writer, sheet_name="cpu_usage_5s", index=False
            )
        else:
            sort_cols = [c for c in ("window_time_local", "window_start_ns", "elapsed_sec") if c in cpu_df.columns]
            sorted_cpu_df = cpu_df.sort_values(sort_cols)
            chart_xy_df = build_chart_xy_dataframe(sorted_cpu_df)
            chart_xy_group_df = build_chart_xy_group_dataframe(sorted_cpu_df)
            format_datetime_columns(sorted_cpu_df).to_excel(writer, sheet_name="cpu_usage_5s", index=False)

        if not chart_xy_df.empty:
            chart_xy_df.to_excel(writer, sheet_name="chart_xy", index=False)
        if not chart_xy_group_df.empty:
            chart_xy_group_df.to_excel(writer, sheet_name="chart_xy_group", index=False)

        chart_gpu_df = pd.DataFrame()
        chart_cpu_gpu_df = pd.DataFrame()
        gpu_sorted = pd.DataFrame()
        if gpu_df is not None and not gpu_df.empty:
            gpu_sorted = gpu_df.sort_values("elapsed_sec")
            gpu_sorted.to_excel(writer, sheet_name="gpu_measure_5s", index=False)
            chart_gpu_df = build_chart_gpu_dataframe(gpu_sorted)
            if not chart_gpu_df.empty:
                chart_gpu_df.to_excel(writer, sheet_name="chart_gpu", index=False)
        elif gpu_df is not None:
            pd.DataFrame({"info": ["未找到 gpuload/gpufreq measure 数据"]}).to_excel(
                writer, sheet_name="gpu_measure_5s", index=False
            )

        chart_cpu_gpu_df = build_chart_cpu_gpu_dataframe(
            sorted_cpu_df,
            gpu_sorted if not gpu_sorted.empty else None,
        )
        if not chart_cpu_gpu_df.empty:
            chart_cpu_gpu_df.to_excel(writer, sheet_name="chart_cpu_gpu", index=False)

    _add_cpu_charts(output_path, chart_xy_df, chart_xy_group_df)
    _add_gpu_charts(output_path, chart_gpu_df)
    _add_cpu_gpu_merged_chart(output_path, chart_cpu_gpu_df)


def _merge_file_meta(cpu_meta: dict, gpu_meta: dict) -> dict:
    by_file: dict[str, dict] = {}
    for row in cpu_meta.get("files", []):
        by_file[row["source_file"]] = dict(row)
    for row in gpu_meta.get("files", []):
        merged = by_file.setdefault(row["source_file"], {"source_file": row["source_file"]})
        merged.update(row)
    return {"files": list(by_file.values())}


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python parse_hitrace_stats.py <trace.db|trace_dir> [output.xlsx]")
        return 1

    input_path = Path(argv[1]).resolve()
    if not input_path.exists():
        print(f"Path not found: {input_path}")
        return 1

    if len(argv) > 2:
        output_path = Path(argv[2]).resolve()
    elif input_path.is_dir():
        output_path = input_path / f"{input_path.name}.stats.xlsx"
    else:
        output_path = input_path.with_suffix(".stats.xlsx")

    print(f"Reading {input_path} ...")
    db_files = parse_db_inputs(input_path)
    intervals, cpu_meta = merge_db_intervals(db_files)
    gpu_samples, gpu_meta = merge_gpu_samples(db_files)
    file_meta = _merge_file_meta(cpu_meta, gpu_meta)
    cpu_df = compute_cpu_usage_5s(intervals)
    gpu_df = compute_gpu_measure_5s(gpu_samples)

    source_label = db_files[0].name if len(db_files) == 1 else f"{len(db_files)} files (time-merged)"
    export_xlsx(cpu_df, file_meta, output_path, source_label=source_label, gpu_df=gpu_df)

    print(f"Done. Output: {output_path}")
    print(f"  db files: {len(db_files)}")
    print(f"  running slices: {len(intervals)}")
    print(f"  cpu_usage_5s rows: {len(cpu_df)}")
    print(f"  gpu measure samples: {len(gpu_samples)}")
    print(f"  gpu_measure_5s rows: {len(gpu_df)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
